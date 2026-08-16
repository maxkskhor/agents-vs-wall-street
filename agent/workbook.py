"""Workbook writer: fills only the yellow forecast cells of the supplied templates.

The Summary sheet layout is fixed: metric rows 7-9 (A=label, B=units,
C=forecast, yellow fill FFFFF7D6), header C6 = fiscal period. We verify the
labels/units/period against challenge/companies.json before writing so a
template drift or metric mix-up fails loudly instead of submitting garbage.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .config import SUBMISSION, Company

YELLOW = "FFFFF7D6"
FORECAST_CELLS = ("C7", "C8", "C9")


class WorkbookError(RuntimeError):
    pass


def write_workbook(company: Company, values: dict[str, float],
                   out_dir: Path | None = None) -> Path:
    """values: metric label -> number (already in workbook units)."""
    out_dir = out_dir or SUBMISSION
    wb = openpyxl.load_workbook(company.template)
    if "Summary" not in wb.sheetnames:
        raise WorkbookError(f"{company.template.name}: missing Summary sheet")
    ws = wb["Summary"]

    if str(ws["C6"].value).strip() != company.period:
        raise WorkbookError(
            f"{company.output_file}: period header {ws['C6'].value!r} != {company.period!r}")

    for row, metric in zip((7, 8, 9), company.metrics):
        label = str(ws[f"A{row}"].value).strip()
        units = str(ws[f"B{row}"].value).strip()
        if label != metric.label or units != metric.units:
            raise WorkbookError(
                f"{company.output_file} row {row}: template says "
                f"({label!r}, {units!r}) but spec says ({metric.label!r}, {metric.units!r})")
        cell = ws[f"C{row}"]
        if cell.fill.start_color.rgb != YELLOW:
            raise WorkbookError(f"{company.output_file}: C{row} is not the yellow forecast cell")
        if metric.label not in values:
            raise WorkbookError(f"{company.output_file}: no value supplied for {metric.label!r}")
        v = float(values[metric.label])
        # money to the unit; EPS and percent keep 4dp — rounding Hays'
        # pence EPS to 2dp can cost most of a score point against the
        # 0.5%-of-actual floor
        cell.value = round(v) if metric.kind == "money" else round(v, 4)

    out_dir.mkdir(exist_ok=True)
    out = out_dir / company.output_file
    wb.save(out)
    return out
